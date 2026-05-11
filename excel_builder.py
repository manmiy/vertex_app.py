import pandas as pd
import re
from datetime import timedelta, date
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import jpholiday

# 祝日判定ツール
def is_holiday_jp(d):
    return d.weekday() >= 5 or jpholiday.is_holiday(d)

# ==========================================
# Excel生成ロジック (完全版)
# ==========================================
def create_filled_excel(df_extracted, sheet1_name="人工集計", sheet2_name="日報明細", target_year=2024):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet1_name
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    data_list = []
    unique_codes = []
    code_to_name = {}
    last_seen_code = ""

    # --- 1. データの前処理と整形 ---
    if df_extracted is not None and not df_extracted.empty:
        for idx, row in df_extracted.iterrows():
            try:
                if len(row) >= 4:
                    m_str = str(row.get('月', '')).replace('月', '').strip()
                    d_str = str(row.get('日', '')).replace('日', '').strip()
                    t_val = str(row.get('時間', '')).strip()

                    code_val = str(row.get('工事コード', '')).strip()
                    code_val = re.sub(r'[^a-zA-Z0-9-]', '', code_val)

                    if not code_val or code_val.lower() in ['nan', 'none']:
                        code_val = last_seen_code
                    else:
                        last_seen_code = code_val

                    if code_val:
                        code_to_name[code_val] = ""

                    content_val = str(row.get('業務内容', '')).strip()
                    t_val = re.sub(r'[^0-9:]', '', t_val)
                    if not t_val: continue

                    m = int(float(m_str))
                    d = int(float(d_str))

                    if code_val and code_val not in unique_codes:
                        unique_codes.append(code_val)

                    break_raw = row.get('休憩', None)
                    if pd.isna(break_raw) or str(break_raw).strip() == "" or break_raw is None:
                        break_val = ""
                    else:
                        try:
                            break_val = float(break_raw)
                        except (ValueError, TypeError):
                            break_val = ""

                    data_list.append({
                        'orig_idx': idx,
                        'm': m, 'd': d, 'break': break_val,
                        'time': t_val, 'code': code_val, 'content': content_val,
                        'category': row.get('区分', "日勤")
                    })
            except Exception as e:
                print(f"Excel前処理エラー (行 {idx}): {e}")

    # --- 2. 1枚目シート（人工集計）の作成 ---
    headers_left = [
        ("B", "月"), ("C", "日"), ("D", "曜日"), ("E", "休・出"), ("F", "摘要"),
        ("G", "始業"), ("H", "終業"), ("I", "時間"), ("J", "休憩"), ("K", "時間"), ("L", "人工")
    ]

    # write_headers関数にremarks_col_idxを追加
    def write_headers(ws, start_row, header_blocks, remarks_col_idx):
        for col, name in headers_left:
            cell = ws[f"{col}{start_row}"]
            cell.value = name
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9, bold=True)
            cell.border = thin_border

        # 実労・計のヘッダー
        ws.merge_cells(f"K{start_row-2}:L{start_row-2}")
        for col in ("K", "L"): ws[f"{col}{start_row-2}"].border = thin_border
        c1 = ws[f"K{start_row-2}"]
        c1.value = "実労"
        c1.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"K{start_row-1}:L{start_row-1}")
        for col in ("K", "L"): ws[f"{col}{start_row-1}"].border = thin_border
        c2 = ws[f"K{start_row-1}"]
        c2.value = "計"
        c2.alignment = Alignment(horizontal="center", vertical="center")

        # 工事コードごとのヘッダー展開
        for block in header_blocks[1:]: # 「実労」ブロック (K/L) はスキップ
            h_code, h_name, col1, col2 = block
            # プロジェクト名/その他名のヘッダー
            ws.merge_cells(f"{col1}{start_row-2}:{col2}{start_row-2}")
            for col_l in (col1, col2): ws[f"{col_l}{start_row-2}"].border = thin_border
            cn = ws[f"{col1}{start_row-2}"]
            cn.value = h_name
            cn.alignment = Alignment(horizontal="center", vertical="center")

            # プロジェクトコード/その他コードのヘッダー
            ws.merge_cells(f"{col1}{start_row-1}:{col2}{start_row-1}")
            for col_l in (col1, col2): ws[f"{col_l}{start_row-1}"].border = thin_border
            cc = ws[f"{col1}{start_row-1}"]
            cc.value = h_code
            cc.alignment = Alignment(horizontal="center", vertical="center")

            # 「時間」と「人工」のヘッダー
            ws[f"{col1}{start_row}"] = "時間"
            ws[f"{col2}{start_row}"] = "人工"
            for col_l in (col1, col2):
                cell = ws[f"{col_l}{start_row}"]
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        
        # 「備考」ヘッダーを追加
        remarks_col_letter = get_column_letter(remarks_col_idx)
        
        # 「備考」セルを3行にわたって結合 (start_row-2からstart_rowまで)
        ws.merge_cells(f"{remarks_col_letter}{start_row-2}:{remarks_col_letter}{start_row}")
        
        cell_remarks = ws[f"{remarks_col_letter}{start_row-2}"] # 結合範囲の左上のセルにアクセス
        cell_remarks.value = "備考"
        cell_remarks.alignment = Alignment(horizontal="center", vertical="center")
        cell_remarks.font = Font(size=9, bold=True)
        
        # 結合された範囲内の全てのセルに罫線を適用して正しく表示されるようにする
        for r_offset in range(3): # start_row-2, start_row-1, start_rowの3行
            ws[f"{remarks_col_letter}{start_row-2 + r_offset}"].border = thin_border

    # 動的なヘッダーブロック（プロジェクトおよび「その他」）の生成
    header_blocks = [("計", "実労", "K", "L")]
    start_col_idx = 13 # プロジェクトの列はM(13)からスタート
    for code in unique_codes:
        col1 = get_column_letter(start_col_idx)
        col2 = get_column_letter(start_col_idx + 1)
        header_blocks.append((code, code_to_name.get(code, ""), col1, col2))
        start_col_idx += 2
    
    # 「その他」ブロックを追加
    other_col1 = get_column_letter(start_col_idx)
    other_col2 = get_column_letter(start_col_idx + 1)
    header_blocks.append(("その他", "", other_col1, other_col2))
    start_col_idx += 2 # start_col_idxは「その他」ブロックの次の列を指す

    # 「備考」列のインデックスを定義
    remarks_col_idx = start_col_idx
    remarks_col_letter = get_column_letter(remarks_col_idx)

    # 列幅の設定 (remarks_col_idxが決定された後に動的な列幅を設定)
    widths = {'A':4, 'B':4, 'C':4, 'D':4, 'E':6, 'F':12, 'G':6, 'H':6, 'I':8, 'J':6, 'K':10, 'L':9}
    widths[remarks_col_letter] = 20 # 新しい「備考」列の幅
    
    # 列幅を適用
    for col, w in widths.items(): ws.column_dimensions[col].width = w

    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    reg_data = [d for d in data_list if d.get('category') == "日勤"]
    hol_data = [d for d in data_list if d.get('category') != "日勤"]

    # fill_rows関数にremarks_col_idxを追加
    def fill_rows(ws, data_rows, start_row_idx, total_label="計", is_continuous=False, remarks_col_idx=None):
        r_idx = start_row_idx
        grouped = {}
        for entry in data_rows:
            md = (entry['m'], entry['d'])
            if md not in grouped: grouped[md] = []
            grouped[md].append(entry)

        def write_single_entry(r_idx, entry, curr_date):
            is_first = entry.get('is_first', True)
            ws[f'B{r_idx}'] = curr_date.month if is_first else ""
            ws[f'C{r_idx}'] = curr_date.day if is_first else ""
            ws[f'D{r_idx}'] = weekdays_jp[curr_date.weekday()] if is_first else ""
            ws[f'E{r_idx}'] = "〇" if is_first else "" 
            ws[f'F{r_idx}'] = "" # 摘要

            entry['final_excel_row'] = r_idx
            ws[f'G{r_idx}'] = '8:00' if is_first else f'=H{r_idx-1}'
            ws[f'H{r_idx}'] = entry['time']
            ws[f'I{r_idx}'] = f'=IF(OR(G{r_idx}="", H{r_idx}=""), "", (H{r_idx}-G{r_idx})*24)'
            b_val = entry.get('break')
            ws[f'J{r_idx}'] = None if b_val == "" else b_val
            
            ws[f'K{r_idx}'] = f'=IF(I{r_idx}="","",I{r_idx}-J{r_idx})'
            ws[f'L{r_idx}'] = f'=IF(K{r_idx}="","",K{r_idx}/7)'

            # プロジェクト固有の時間/人工列を決定
            # unique_codesにない場合、「その他」列へ
            col_t_idx = 13 + unique_codes.index(entry['code']) * 2 if entry.get('code') in unique_codes else remarks_col_idx - 2

            ws[f'{get_column_letter(col_t_idx)}{r_idx}'] = f'=K{r_idx}'
            ws[f'{get_column_letter(col_t_idx+1)}{r_idx}'] = f'=L{r_idx}'

            # 「備考」列を含む全ての列に罫線と配置を適用
            for col_idx_loop in range(2, remarks_col_idx + 1):
                cell = ws[f'{get_column_letter(col_idx_loop)}{r_idx}']
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx_loop in [7, 8]: cell.number_format = '[h]:mm'
                elif col_idx_loop in [9, 10, 11]: cell.number_format = '0.00'
                elif col_idx_loop == 12: cell.number_format = '0.0000'
                # 「備考」列には数値書式を適用しない
                elif col_idx_loop >= 13 and col_idx_loop < remarks_col_idx:
                    cell.number_format = '0.00' if (col_idx_loop % 2 != 0) else '0.0000'

        if is_continuous:
            if data_rows:
                sample = data_rows[0]
                sd = date(target_year, sample['m'], sample['d'])
            else:
                sd = date.today()

            if sd.day >= 21:
                period_start = date(sd.year, sd.month, 21)
            else:
                prev_m = 12 if sd.month == 1 else sd.month - 1
                prev_y = sd.year - 1 if sd.month == 1 else sd.year
                period_start = date(prev_y, prev_m, 21)

            next_m = period_start.month % 12 + 1
            next_y = period_start.year + (1 if period_start.month == 12 else 0)
            period_end = date(next_y, next_m, 20)

            curr = period_start
            while curr <= period_end:
                md = (curr.month, curr.day)
                entries = grouped.get(md, [])

                if not entries:
                    ws[f'B{r_idx}'] = curr.month
                    ws[f'C{r_idx}'] = curr.day
                    ws[f'D{r_idx}'] = weekdays_jp[curr.weekday()]
                    ws[f'E{r_idx}'] = "×" 
                    ws[f'F{r_idx}'] = ""  
                    # 「備考」列を含む全ての列に罫線を適用
                    for col_idx_loop in range(7, remarks_col_idx + 1):
                        cell = ws[f'{get_column_letter(col_idx_loop)}{r_idx}']
                        cell.value = "" if col_idx_loop > 8 else cell.value 
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    r_idx += 1
                else:
                    for t_idx, entry in enumerate(entries):
                        entry['is_first'] = (t_idx == 0)
                        write_single_entry(r_idx, entry, curr)
                        r_idx += 1
                curr += timedelta(days=1)
        else:
            for md, entries in grouped.items():
                for t_idx, entry in enumerate(entries):
                    entry['is_first'] = (t_idx == 0)
                    try: curr_date = date(target_year, entry['m'], entry['d'])
                    except: curr_date = date(target_year, 1, 1)
                    write_single_entry(r_idx, entry, curr_date)
                    r_idx += 1

        if r_idx > start_row_idx:
            ws.merge_cells(f"B{r_idx}:D{r_idx}")
            ws[f"B{r_idx}"] = total_label
            ws[f"B{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"E{r_idx}"] = f'=COUNTIF(E{start_row_idx}:E{r_idx-1}, "〇")'
            ws[f"E{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
            
            # 合計を計算する列 (「備考」列の手前まで)
            for col_idx_loop in range(9, remarks_col_idx):
                col_let = get_column_letter(col_idx_loop)
                ws[f"{col_let}{r_idx}"] = f"=SUM({col_let}{start_row_idx}:{col_let}{r_idx-1})"
                if col_idx_loop == 12 or (col_idx_loop >= 14 and col_idx_loop % 2 == 0):
                    ws[f"{col_let}{r_idx}"].number_format = '0.0000'
                else:
                    ws[f"{col_let}{r_idx}"].number_format = '0.00'
            
            # 合計行の罫線とフォントを「備考」列を含む全ての列に適用
            for col_idx_loop in range(2, remarks_col_idx + 1):
                cell = ws[f'{get_column_letter(col_idx_loop)}{r_idx}']
                cell.border = thin_border
                cell.font = Font(bold=True)
            r_idx += 1
        return r_idx

    write_headers(ws, 4, header_blocks, remarks_col_idx) # remarks_col_idxを渡す
    next_row = fill_rows(ws, reg_data, 5, total_label="日勤計", is_continuous=True, remarks_col_idx=remarks_col_idx) # remarks_col_idxを渡す
    
    if hol_data:
        next_row += 3
        ws.merge_cells(f"A{next_row-1}:A{next_row + len(hol_data) + (1 if len(hol_data)>0 else 0)}")
        cell_a = ws[f"A{next_row-1}"]
        cell_a.value = "残業・休日出勤"
        cell_a.alignment = Alignment(horizontal="center", vertical="center", textRotation=255)
        cell_a.border = thin_border
        write_headers(ws, next_row, header_blocks, remarks_col_idx) # remarks_col_idxを渡す
        next_row = fill_rows(ws, hol_data, next_row + 1, total_label="残業計", is_continuous=False, remarks_col_idx=remarks_col_idx) # remarks_col_idxを渡す

    sheet1_row_mapping = {}
    for d in data_list:
        if 'orig_idx' in d and 'final_excel_row' in d:
            sheet1_row_mapping[d['orig_idx']] = d['final_excel_row']

    # --- 3. 2枚目のシート(日報明細)を作成 ---
    ws2 = wb.create_sheet(title=sheet2_name)
    headers_s2 = ["月", "日", "始業", "終業", "時間", "休憩", "実労時間", "工事コード", "現場名", "業務内容"]
    
    for col_idx, h in enumerate(headers_s2, start=1):
        c = ws2.cell(row=2, column=col_idx, value=h)
        c.font = Font(bold=True)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")

    if df_extracted is not None and not df_extracted.empty:
        r_idx = 3 
        prev_m, prev_d = None, None
        prev_m_time, prev_d_time = None, None
        
        for idx, row in df_extracted.fillna('').iterrows():
            m_str = str(row.get('月', '')).replace('.0', '').strip()
            d_str = str(row.get('日', '')).replace('.0', '').strip()

            try: m_num = int(float(m_str))
            except ValueError: m_num = m_str

            try: d_num = int(float(d_str))
            except ValueError: d_num = d_str

            t_val = str(row.get('時間', '')).strip()
            t_val = re.sub(r'[^0-9:]', '', t_val)

            if m_num == prev_m_time and d_num == prev_d_time and m_num != "":
                is_first = False
            else:
                is_first = True
            prev_m_time = m_num
            prev_d_time = d_num

            break_raw = row.get('休憩', '')
            if pd.isna(break_raw) or str(break_raw).strip() == "" or break_raw is None:
                b_val_ext = None
            else:
                try: b_val_ext = float(break_raw)
                except (ValueError, TypeError): b_val_ext = None

            c_val = str(row.get('工事コード', ''))
            p_val = str(row.get('現場名', ''))
            b_val = str(row.get('業務内容', ''))

            if m_num == prev_m and d_num == prev_d and m_num != "":
                out_m, out_d = "", ""
            else:
                out_m, out_d = m_num, d_num
            if m_num != "":
                prev_m, prev_d = m_num, d_num

            ws2[f'A{r_idx}'] = out_m
            ws2[f'B{r_idx}'] = out_d
            ws2[f'C{r_idx}'] = '8:00' if is_first else f'=D{r_idx-1}'
            ws2[f'D{r_idx}'] = t_val
            ws2[f'E{r_idx}'] = f'=IF(OR(C{r_idx}="", D{r_idx}=""), "", (D{r_idx}-C{r_idx})*24)'
            
            if idx in sheet1_row_mapping:
                s1_r = sheet1_row_mapping[idx]
                ws2[f'F{r_idx}'] = f'=IF(\'{sheet1_name}\'!J{s1_r}="","",\'{sheet1_name}\'!J{s1_r})'
            else:
                ws2[f'F{r_idx}'] = b_val_ext
                
            ws2[f'G{r_idx}'] = f'=IF(E{r_idx}="","",E{r_idx}-IF(F{r_idx}="",0,F{r_idx}))'
            ws2[f'H{r_idx}'] = c_val
            ws2[f'I{r_idx}'] = p_val
            ws2[f'J{r_idx}'] = b_val

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                c = ws2[f'{col_letter}{r_idx}']
                c.border = thin_border
                if col_letter in ['I', 'J']: c.alignment = Alignment(wrapText=True, vertical="center")
                else: c.alignment = Alignment(horizontal="center", vertical="center")

                # Sheet2の数値書式を修正
                current_cell_col_idx = c.column
                if current_cell_col_idx in [3, 4]: # C, D
                    c.number_format = '[h]:mm'
                elif current_cell_col_idx in [5, 6, 7]: # E, F, G
                    c.number_format = '0.00'

            r_idx += 1

        ws2.column_dimensions['A'].width = 6
        ws2.column_dimensions['B'].width = 6
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 10
        ws2.column_dimensions['G'].width = 10
        ws2.column_dimensions['H'].width = 15
        ws2.column_dimensions['I'].width = 20
        ws2.column_dimensions['J'].width = 40

    return wb